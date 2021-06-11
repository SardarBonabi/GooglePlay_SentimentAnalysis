import openpyxl
# for scraping app info and reviews from Google Play
from google_play_scraper import app, Sort, reviews

from pymongo import MongoClient

# for keeping track of timing
import datetime as dt
from tzlocal import get_localzone

# for building in wait times
import random
import time

# Set up Mongo client
client = MongoClient(host='localhost', port=27017)

## Database for projectx`
app_proj_db = client['app_proj_db6']

## Set up new collection within project db for app info
info_collection = app_proj_db['info_collection']

## Set up new collection within project db for app reviews
review_collection = app_proj_db['review_collection']


# load excel with its path
wrkbk = openpyxl.load_workbook("App Names and IDs.xlsx")
sh = wrkbk.active

# iterate through excel and get app ids
app_ids = []
for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
# for row in sh.iter_rows(min_row=2, min_col=2, max_row=10, max_col=2):
    for cell in row:
        app_ids.append(cell.value)

# iterate through excel and get app names
app_names = []
for row in sh.iter_rows(min_row=2, min_col=1, max_row=sh.max_row, max_col=1):
# for row in sh.iter_rows(min_row=2, min_col=1, max_row=10, max_col=1):
    for cell in row:
        app_names.append(cell.value)








## Loop through app IDs to get app info
app_info = []
for i in app_ids:
    print("full info of", i, " is scraped")
    info = app(i)
    del info['comments']
    app_info.append(info)

## Pretty print the data for the first app
# pprint(app_info[0])

## Insert app details into info_collection
info_collection.insert_many(app_info)


## Loop through apps to get reviews
for app_name, app_id in zip(app_names, app_ids):

    # Get start time
    start = dt.datetime.now(tz=get_localzone())
    fmt = "%m/%d/%y - %T %p"

    # Print starting output for app
    # print('---' * 20)
    # print('---' * 20)
    print(f'***** {app_name} started at {start.strftime(fmt)}')
    print()

    # Empty list for storing reviews
    app_reviews = []

    # Number of reviews to scrape per batch
    count = 200

    # To keep track of how many batches have been completed
    batch_num = 0

    # Retrieve reviews (and continuation_token) with reviews function
    rvws, token = reviews(
        app_id,  # found in app's url
        lang='en',  # defaults to 'en'
        country='us',  # defaults to 'us'
        sort=Sort.MOST_RELEVANT,  # start with most recent
        count=count  # batch size
    )

    # For each review obtained
    for r in rvws:
        r['app_name'] = app_name  # add key for app's name
        r['app_id'] = app_id  # add key for app's id

    # Add the list of review dicts to overall list
    app_reviews.extend(rvws)

    # Increase batch count by one
    batch_num += 1
    print(f'Batch {batch_num} completed.')

    # Wait 1 to 5 seconds to start next batch
    time.sleep(random.randint(5, 20))

    # Append review IDs to list prior to starting next batch
    pre_review_ids = []
    for rvw in app_reviews:
        pre_review_ids.append(rvw['reviewId'])

    # Loop through at most max number of batches
    # for batch in range(4999):
    for batch in range(49):
        rvws, token = reviews(  # store continuation_token
            app_id,
            lang='en',
            country='us',
            sort=Sort.NEWEST,
            count=count,
            # using token obtained from previous batch
            continuation_token=token
        )

        # Append unique review IDs from current batch to new list
        new_review_ids = []
        for r in rvws:
            new_review_ids.append(r['reviewId'])

            # And add keys for name and id to ea review dict
            r['app_name'] = app_name  # add key for app's name
            r['app_id'] = app_id  # add key for app's id

        # Add the list of review dicts to main app_reviews list
        app_reviews.extend(rvws)

        # Increase batch count by one
        batch_num += 1

        # Break loop and stop scraping for current app if most recent batch
        # did not add any unique reviews
        all_review_ids = pre_review_ids + new_review_ids
        if len(set(pre_review_ids)) == len(set(all_review_ids)):
            print(f'No reviews left to scrape. Completed {batch_num} batches.\n')
            break

        # all_review_ids becomes pre_review_ids to check against
        # for next batch
        pre_review_ids = all_review_ids

        # At every 100th batch
        if batch_num % 100 == 0:
            # print update on number of batches
            print(f'Batch {batch_num} completed.')

            # insert reviews into collection
            review_collection.insert_many(app_reviews)

            # print update about num reviews inserted
            store_time = dt.datetime.now(tz=get_localzone())
            print(f"""
            Successfully inserted {len(app_reviews)} {app_name} 
            reviews into collection at {store_time.strftime(fmt)}.\n
            """)

            # empty our list for next round of 100 batches
            app_reviews = []

        # Wait 1 to 5 seconds to start next batch
        time.sleep(random.randint(5, 15))

    # Print update when max number of batches has been reached
    # OR when last batch didn't add any unique reviews
    print(f'Done scraping {app_name}.')
    print(f'Scraped a total of {len(set(pre_review_ids))} unique reviews.\n')

    if len(app_reviews) > 0:
    # Insert remaining reviews into collection
        review_collection.insert_many(app_reviews)

    # Get end time
    end = dt.datetime.now(tz=get_localzone())

    # Print ending output for app
    print(f"""
    Successfully inserted all {app_name} reviews into collection
    at {end.strftime(fmt)}.\n
    """)
    print(f'Time elapsed for {app_name}: {end - start}')
    print('---' * 20)
    print('---' * 20)
    print('\n')

    # Wait 1 to 5 seconds to start scraping next app
    time.sleep(random.randint(5, 10))

import pandas as pd
from pymongo import MongoClient
import json

path = "C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\4. Project\\1. Phase 1\\Comments and Apps\\"

## Set up Mongo client
client = MongoClient(host='localhost', port=27017)
db = client['app_proj_db5']
collection = db['review_collection']
print("Connection to Database Successful!")

# Get Applications' ids from database
app_ids = collection.distinct('app_id')

AppNumber = 0

for app_id in app_ids:
    print("App " + str(AppNumber) + " Started")
    reviewsList = []
    reviews = collection.find({"app_id": app_id}, {"_id": 0, "content": 1}).limit(2000)

    for review in reviews:
        review_json = json.dumps(review)
        review = json.loads(review_json)
        reviewsList.append(review["content"])

    content = reviewsList
    sentiment = [None] * len(content)

    data = {'sentiment': sentiment,
            'content': content
            }

    df = pd.DataFrame(data, columns=['sentiment', 'content'])
    df.to_csv(path + "Comments_App" + str(AppNumber) + " - " + app_id + ".csv", index=False, header=True)
    print("App " + str(AppNumber) + " Finished")
    AppNumber += 1
    print("---------------------------------------------------")

print("********* All Done! *********")

import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.linear_model import LogisticRegression
from os import listdir
from os.path import isfile, join
from sklearn.tree import DecisionTreeClassifier

print("Application Started!")

# Read Train Data
df_emotions = pd.read_csv(
    "C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\1. Lectures\\6. 5.12.2021\\emotions.csv")
train = df_emotions

print("Train File Reading Done")

# Used Vectorizer
vectorizer = CountVectorizer(ngram_range=(1, 1))
vectorizer.fit(train.content)

model = DecisionTreeClassifier()
model.fit(vectorizer.transform(train.content), train.sentiment)

print("Training the Model Done")
print("----------------------------------------")

mypath = "C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\4. Project\\1. Phase 1\\Comments and Apps\\"
fileNames = [f for f in listdir(mypath) if isfile(join(mypath, f))]

# Remove Blank lines
correctedPath = "C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\4. Project\\1. Phase 1\\Comments and Apps Corrected\\"

print("Verifying CSV files...")

for fileName in fileNames:
    data = pd.read_csv(mypath + fileName, skip_blank_lines=True)
    data.dropna(how="all", inplace=True)
    data.to_csv(correctedPath + fileName, index=False, header=True)

print("Verification Done")
print("--------------------------------")

# Make Predictions
i = 0
for fileName in fileNames:
    print("App " + str(i) + " Started")
    test = pd.read_csv(
        "C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\4. Project\\1. Phase 1\\Comments and Apps Corrected\\" + fileName)

    prediction = model.predict(vectorizer.transform(test.content))
    print("App " + str(i) + " Prediction Done")

    data = {'sentiment': prediction, 'content': test["content"]}

    df = pd.DataFrame(data, columns=['sentiment', 'content'])
    savepath = "C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\4. Project\\1. Phase 1\\Labeled Comments\\"
    df.to_csv(savepath + fileName, index=False, header=True)
    print("App " + str(i) + " Done")
    i += 1
    print("----------------------------------------")

print("********* All Done! *********")


from wordcloud import WordCloud, STOPWORDS, ImageColorGenerator

sentiments = ""
for data in test['sentiment']:
    sentiments = sentiments + " " + data

# Generate a word cloud image
wordcloud = WordCloud(background_color="white").generate(sentiments)

# Display the generated image:
# the matplotlib way:
plt.imshow(wordcloud, interpolation='bilinear')
plt.axis("off")
plt.show()


df_emotions = pd.read_csv(
    "C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\1. Lectures\\6. 5.12.2021\\emotions.csv")
train = df_emotions

test = test = pd.read_csv("C:\\Users\\Sardar\\Desktop\\UCI\\Education\\3. Spring 2021 Q\\1. NLP\\4. Project\\1. Phase 1\\Labeled Comments\\Comments_App20 - buttocksworkout.hipsworkouts.forwomen.legworkout.csv")

test['sentiment'].value_counts().plot(kind='bar')


def plot_confusion_matrix(preds, labels):
  class_labels = np.unique(df_emotions.sentiment)
  class_size = len(class_labels)
  cnf_mat = confusion_matrix(labels, preds, class_labels)                            #Computes confusion_matrix
  cnf_mat = cnf_mat.astype('float') / (cnf_mat.sum(axis=1)[:, np.newaxis]+1)
  plt.imshow(cnf_mat, interpolation='nearest', cmap=plt.cm.Blues)
  #plt.xticks(np.arange(class_size), np.arange(1, class_size + 1), class_labels)
  plt.xticks(np.arange(class_size), labels=class_labels, rotation='vertical')
  plt.yticks(np.arange(class_size), labels=class_labels)
  #plt.yticks(np.arange(class_size), np.arange(1, class_size + 1), class_labels)
  plt.title('Confusion matrix of the classifier')
  plt.xlabel('True Label')
  plt.ylabel('Predicted Label')
  plt.title('Confusion Matrix')
  plt.colorbar()
  plt.show()


from sklearn.metrics import confusion_matrix, precision_score, precision_recall_curve, recall_score, f1_score, accuracy_score

tv = CountVectorizer(ngram_range=(1,1))
tv.fit(train.content)

modelELR = DecisionTreeClassifier()
modelELR.fit(tv.transform(train.content), train.sentiment)
elr_pred = modelELR.predict(tv.transform(test.content))

print("Confusion Matrix:\n", confusion_matrix(test.sentiment, elr_pred))
print("F1 score:", f1_score(test.sentiment, elr_pred, average='micro'))


import numpy as np
import matplotlib.pyplot as plt

plot_confusion_matrix(test.sentiment, elr_pred)

